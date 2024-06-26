import speech_recognition as sr
import openpyxl


rec = sr.Recognizer()
mic = sr.Microphone()


def create_text_from_audio(audio_file: str) -> str:
    """Converts audio file to text."""
    try:
        billing_audio = sr.AudioFile(audio_file)
        with billing_audio as source:
            rec.adjust_for_ambient_noise(source)
            audio = rec.record(source)
        text = rec.recognize_google(audio)
        print("Google Speech Recognition converted your recording to the following text\n\n" + text)
    except sr.UnknownValueError:
        print("Google Speech Recognition could not understand audio")
        text = "Unrecognized"
    except sr.RequestError as e:
        print("Could not request results from Google Speech Recognition service; {0}".format(e))
        text = "Error"
    return text


def sort_entries(entry_list: list) -> tuple[list, list]:
    """Checks if an entry is valid and appends it to the correct list."""
    check_again = list()
    cleared = list()
    for entry in entry_list:
        if len(entry.split()) != 3:
            check_again.append(entry)
        else:
            valid_entry = entry.split()
            cleared.append(valid_entry)
    return cleared, check_again


def convert_month(m: str) -> str:
    """Converts the word of a month into the according string of ciphers."""
    months = {"january": "01", "february": "02",  "march": "03", "april": "04", "may": "05",
              "june": "06", "july": "07", "august": "08", "september": "09", "october": "10",
              "november": "11", "december": "12"}
    current_month = m
    try:
        current_month = months[m.lower()]
    except:
        print("Error: month not recognized.")
    finally:
        return current_month


def convert_amount(a: str) -> float | str:
    """Converts a string to a float and returns it if successful."""
    amount = a.strip()
    try:
        amount = float(a)
    except:
        print("Error: not a number")
    finally:
        return amount


def create_full_date(d: str, m: str, y: str) -> str:
    """Concatenates year, month and day to one string."""
    return y + convert_month(m) + d

print()
billing_data = create_text_from_audio("test3.wav")
bd_listed = billing_data.split("next")
valid_entries, invalid_entries = sort_entries(bd_listed)

current_year = "2024"
current_month = invalid_entries.pop(0).strip()

print("\nCreating spreadsheet for:", current_month)
print("\nValid entries:")
for entry in valid_entries:
    print("\t", entry)
print("\nNo invalid entries" if not invalid_entries else f"\nEntries to be checked manually in Excel file: {invalid_entries}")

wb = openpyxl.load_workbook("billing.xlsx")
ws = wb.create_sheet(current_month)

ws["A1"] = "Date"
ws["B1"] = "Category"
ws["C1"] = "Amount"

for entry in valid_entries:
    row_number = ws.max_row + 1
    ws["A" + str(row_number)] = create_full_date(entry[0].strip(), current_month, current_year)
    ws["B" + str(row_number)] = entry[1].strip()
    ws["C" + str(row_number)] = convert_amount(entry[2])

for entry in invalid_entries:
    row_number = ws.max_row + 1
    ws["E" + str(row_number)] = entry

row_number = ws.max_row + 2
ws["B" + str(row_number)] = "Total"
ws["C" + str(row_number)] = "=SUM(C2:C" + str(row_number-1) + ")"

wb.save("billing.xlsx")
print("\nData saved to file.")
wb.close()
